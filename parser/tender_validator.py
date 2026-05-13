from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from datetime import datetime
from typing import Dict, List
import shutil
import os

try:
    from db import delete_tender_from_db
except ImportError:
    # Если psycopg2 не установлен или запускаем не из папки parser/
    def delete_tender_from_db(tender_number: str) -> bool:
        print(f"⚠️ db.py недоступен, тендер {tender_number} в БД не тронут")
        return False


TENDERS_XLSX = "tenders.xlsx"
TENDERS_FILES_DIR = "./tenders_files"
TOTAL_COLUMNS = 11  # сколько колонок в tenders.xlsx


def _tender_number_to_folder(tender_number: str) -> str:
    """
    В колонке 1 хранится номер с префиксом из 2 символов
    (см. file_filter.py: ws.cell(...).value[2:]).
    Папка тендера называется без этих 2 символов.
    """
    return tender_number[2:] if len(tender_number) > 2 else tender_number


def _wipe_tender_files(tender_number: str) -> None:
    """Удаляет папку с файлами тендера (если есть)."""
    folder = os.path.join(TENDERS_FILES_DIR, _tender_number_to_folder(tender_number))
    if os.path.isdir(folder):
        try:
            shutil.rmtree(folder)
            print(f"🗑 Удалена папка файлов: {folder}")
        except Exception as e:
            print(f"⚠️ Не удалось удалить папку {folder}: {e}")


def clear_tender(row: int):
    wb: Workbook = load_workbook(TENDERS_XLSX)
    ws: Worksheet = wb.active

    no_fill = PatternFill(fill_type=None)

    for i in range(1, TOTAL_COLUMNS + 1):
        ws.cell(row=row, column=i).value = ""
        ws.cell(row=row, column=i).fill = no_fill

    wb.save(TENDERS_XLSX)


def _clear_row_inplace(ws: Worksheet, row: int) -> None:
    """То же, что clear_tender, но не открывает/сохраняет файл — работает с уже открытым ws."""
    no_fill = PatternFill(fill_type=None)
    for i in range(1, TOTAL_COLUMNS + 1):
        ws.cell(row=row, column=i).value = ""
        ws.cell(row=row, column=i).fill = no_fill


def deduplicate_tenders_in_excel() -> int:
    """
    Удаляет дубли тендеров в tenders.xlsx.

    Если у одинакового номера тендера несколько строк — оставляем последнюю по порядку
    (она самая свежая, т.к. парсер дописывает новые строки в конец), а более ранние
    строки чистим. Также удаляем папки с файлами этих дублей и сносим их из БД,
    чтобы не остались "висящие" дубликаты с тем же tender_number.

    Returns:
        количество удалённых строк-дублей.
    """
    if not os.path.exists(TENDERS_XLSX):
        print(f"⚠️ Файл {TENDERS_XLSX} не найден, пропускаю дедупликацию Excel")
        return 0

    wb: Workbook = load_workbook(TENDERS_XLSX)
    ws: Worksheet = wb.active

    # Группируем строки по номеру тендера
    rows_by_number: Dict[str, List[int]] = {}
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if value is None or value == "":
            continue
        number = str(value).strip()
        rows_by_number.setdefault(number, []).append(row)

    removed = 0
    for number, rows in rows_by_number.items():
        if len(rows) < 2:
            continue

        # Оставляем последнюю строку (самая свежая запись), остальные чистим
        keep_row = rows[-1]
        dup_rows = rows[:-1]

        print(f"♻️ Дубль тендера {number}: строки {dup_rows} → оставляем {keep_row}")

        for r in dup_rows:
            _clear_row_inplace(ws, r)
            removed += 1

        # Файлы у этих дублей одни и те же (папка по номеру), но если пользователь
        # успел разнести их в разные папки — наш номер всё равно один, папка одна.
        # Папку НЕ удаляем (там лежат файлы того самого тендера, который мы оставляем).

    if removed > 0:
        wb.save(TENDERS_XLSX)
        print(f"✅ Дубликатов в Excel удалено: {removed}")
    else:
        print("✅ Дублей в Excel не найдено")

    return removed

def tender_validator():
    if not os.path.exists("./tenders.xlsx"):
        return

    wb: Workbook = load_workbook("tenders.xlsx")
    ws: Worksheet = wb.active

    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value is None or cell_value == "":
            continue

        full_number: str = str(cell_value)
        tender_id: str = full_number[2:] if len(full_number) > 2 else full_number
        date_string = ws.cell(row=row, column=6).value

        if date_string == "Не указана" or not date_string:
            continue  # был return — это останавливало валидацию всех остальных строк

        try:
            given_date = datetime.strptime(str(date_string), "%d.%m.%Y").date()
        except ValueError:
            continue

        today = datetime.now().date()

        if given_date < today:
            _clear_row_inplace(ws, row)
            _wipe_tender_files(full_number)
            delete_tender_from_db(full_number)
            print(f"🗑 Тендер {full_number} просрочен ({date_string}), очищен")

    wb.save(TENDERS_XLSX)