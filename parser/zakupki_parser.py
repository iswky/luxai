from typing import Dict, Set, Optional, Any
from bs4 import BeautifulSoup
from datetime import datetime
import traceback
import requests
import time
import csv
import os
import re

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
    print("openpyxl loaded successfully")
except ImportError as e:
    OPENPYXL_AVAILABLE = False
    print(f"openpyxl not installed: {e}")

# tender parser with extended data
class TenderParser:

    red_fill: PatternFill
    excel_file: str
    csv_file: str
    headers: Dict[str, str]
    base_url: str
    base_params: Dict[str, str]
    existing_numbers: Set[str]

    def __init__(self, excel_file='tenders.xlsx'):
        self.red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.excel_file = excel_file
        self.csv_file = excel_file.replace('.xlsx', '.csv')
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
        }
        
        self.base_url = 'https://zakupki.gov.ru/epz/order/extendedsearch/results.html'
        
        self.base_params = self.load_filters_from_file("./params.txt")
        
        # load existing numbers at startup (one time)
        self.existing_numbers = self.load_existing_tenders()
        
        print("\n" + "="*60)
        print("PARSER STARTED")
        print("="*60)
    
    # reads filters from a txt file in the key=value format
    def load_filters_from_file(self, file_path: str) -> Dict[str, str]:
        params: Dict[str, str] = {}
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    # skip empty lines and comments
                    if not line or line.startswith('#'):
                        continue
                    
                    # divide by the first equal sign
                    if '=' in line:
                        key, value = line.split('=', 1)
                        params[key.strip()] = value.strip()
                        
        except FileNotFoundError:
            print(f"File {file_path} not found")
        except Exception as e:
            print(f"Error reading file: {e}")
        
        return params

    # grabs existing tender numbers from excel
    def load_existing_tenders(self):
        existing_numbers = set()
        
        if OPENPYXL_AVAILABLE and os.path.exists(self.excel_file):
            try:
                wb = load_workbook(self.excel_file)
                ws = wb.active
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        existing_numbers.add(str(row[0]).strip())
                
                print(f"Existing tenders loaded: {len(existing_numbers)}")
            except Exception as e:
                print(f"Error loading: {e}")
        else:
            print("File not found or empty. All tenders will be added.")
        
        return existing_numbers
    
    # stashes new tenders (only those that don't exist yet)
    def save_tenders(self, new_tenders):
        if not new_tenders:
            return 0
        
        # we filter only new tenders
        truly_new = {}
        for number, info in new_tenders.items():
            if number not in self.existing_numbers:
                truly_new[number] = info
                self.existing_numbers.add(number)  # add to cache
        
        if not truly_new:
            print(f"No new tenders to save")
            return 0
        
        if OPENPYXL_AVAILABLE:
            result = self._append_to_excel(truly_new)
        else:
            result = self._append_to_csv(truly_new)
        
        return result
    
    # finds all empty rows in a table.spits out a list of line numbers where the first column is empty.
    def _find_empty_rows(self, ws):
        empty_rows = []
        max_row = ws.max_row
        
        # if the table has only headers, there are no empty rows
        if max_row <= 1:
            return []
        
        # collecting all line numbers with an empty first column
        for row in range(2, max_row + 1):
            if ws.cell(row=row, column=1).value is None:
                empty_rows.append(row)
            else:
                # checking if the line is empty (all columns are empty)
                is_completely_empty = True
                for col in range(1, 12):  # checking all 11 columns
                    if ws.cell(row=row, column=col).value is not None:
                        is_completely_empty = False
                        break
                if is_completely_empty:
                    empty_rows.append(row)
        
        return sorted(empty_rows)  # sort for sequential filling

    # adds new tenders in excel, filling in the blank lines at the beginning
    def _append_to_excel(self, new_tenders):
        try:
            # open or create a file
            if os.path.exists(self.excel_file):
                wb = load_workbook(self.excel_file)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Тендеры"
                headers = ['Номер тендера', 'Название/Объект закупки', 'Ссылка', 
                          'Цена (руб)', 'ФЗ', 'Окончание подачи заявок', 
                          'Заказчик', 'Дата добавления', 'Город', 'Файлы скачаны', 'Файлы отфильтрованы']
                ws.append(headers)
            
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 70
            ws.column_dimensions['C'].width = 60
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 50
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 30
            ws.column_dimensions['J'].width = 20
            ws.column_dimensions['K'].width = 30

            # finding empty lines
            empty_rows = self._find_empty_rows(ws)
            
            # creating a queue of empty lines
            from collections import deque
            empty_rows_queue = deque(empty_rows)
            
            # preparing data to add
            added_count = 0
            next_new_row = ws.max_row + 1
            
            for number, info in new_tenders.items():
                # take the next empty line or create a new one
                if empty_rows_queue:
                    target_row = empty_rows_queue.popleft()
                    location = f"пустую строку {target_row}"
                else:
                    target_row = next_new_row
                    next_new_row += 1
                    location = f"новую строку {target_row}"
                
                # recording data
                ws.cell(row=target_row, column=1, value=number)
                ws.cell(row=target_row, column=2, value=info['name'])
                ws.cell(row=target_row, column=3, value=info['url'])
                ws.cell(row=target_row, column=4, value=info.get('price', 'Не указана'))
                ws.cell(row=target_row, column=5, value=info.get('law', 'Не указан'))
                ws.cell(row=target_row, column=6, value=info.get('end_date', 'Не указана'))
                ws.cell(row=target_row, column=7, value=info.get('customer', 'Не указан'))
                ws.cell(row=target_row, column=8, value=info['first_seen'])
                ws.cell(row=target_row, column=9, value=info.get('city', 'Не указан'))
                ws.cell(row=target_row, column=10, value='False')
                ws.cell(row=target_row, column=10).fill = self.red_fill
                ws.cell(row=target_row, column=11, value='False')
                ws.cell(row=target_row, column=11).fill = self.red_fill
                
                added_count += 1
                print(f"Tender {number} added to {location}")
            
            wb.save(self.excel_file)
            print(f"Added {added_count} tenders (empty rows used: {len(empty_rows[:added_count])})")
            return added_count
            
        except Exception as e:
            print(f"Save error: {e}")
            traceback.print_exc()
            return 0
    
    # stashes to csv with columns city and files downloaded
    def _append_to_csv(self, new_tenders):
        try:
            file_exists = os.path.exists(self.csv_file)
            
            with open(self.csv_file, 'a', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                if not file_exists:
                    # headings with new columns
                    writer.writerow(['number', 'name', 'url', 'price', 'law', 'end_date', 
                                    'customer', 'first_seen', 'city', 'files_downloaded'])
                
                for number, info in new_tenders.items():
                    writer.writerow([
                        number, 
                        info['name'], 
                        info['url'],
                        info.get('price', 'Не указана'),
                        info.get('law', 'Не указан'),
                        info.get('end_date', 'Не указана'),
                        info.get('customer', 'Не указан'),
                        info['first_seen'],
                        info.get('city', 'Не указан'),  # city (still a stub, later you can parse it)
                        'False'                         # files downloaded = false
                    ])
            
            print(f"Saved {len(new_tenders)} new tenders to {self.csv_file}")
            return len(new_tenders)
        except Exception as e:
            print(f"Error: {e}")
            return 0
    
    # chews through the page
    def parse_page(self, page_number: int = 1):

        # params = self.base_params.copy()
        params: Dict[str, str] = {}
        params['pageNumber'] = str(page_number)
        
        url = "https://zakupki.gov.ru/epz/order/extendedsearch/results.html?morphology=on&search-filter=%D0%94%D0%B0%D1%82%D0%B5+%D1%80%D0%B0%D0%B7%D0%BC%D0%B5%D1%89%D0%B5%D0%BD%D0%B8%D1%8F&sortDirection=false&recordsPerPage=_10&showLotsInfoHidden=false&sortBy=UPDATE_DATE&fz44=on&fz223=on&af=on&currencyIdGeneral=-1&okpd2Ids=8889332%2C8889331%2C8873906&okpd2IdsCodes=32.99.53.190%2C32.99.53.130%2C26"

        try:
            print(f"Loading page {page_number}...")
            response = requests.get(url, headers=self.headers, params=params, timeout=15)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            return self.extract_tenders(soup), soup
        except Exception as e:
            print(f"Page error {page_number}: {e}")
            return [], None
    
    # retrieves tenders with enhanced data
    def extract_tenders(self, soup):
        tenders = {}
        current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        blocks = soup.find_all('div', class_='search-registry-entry-block')
        print(f"Blocks found: {len(blocks)}")
        
        for block in blocks:
            try:
                # number and link
                number_div = block.find('div', class_='registry-entry__header-mid__number')
                if not number_div:
                    continue
                
                link = number_div.find('a')
                if not link:
                    continue
                
                number = link.text.strip()
                url = link.get('href')
                
                # name
                name_div = block.find('div', class_='registry-entry__body-value')
                name = name_div.text.strip() if name_div else 'Название не указано'
                if len(name) > 200:
                    name = name[:197] + '...'
                
                # customer
                customer = 'Не указан'
                customer_block = block.find('div', class_='registry-entry__body-block')
                if customer_block:
                    customer_title = customer_block.find('div', class_='registry-entry__body-title', string=re.compile(r'Заказчик|Организация'))
                    if customer_title:
                        customer_value = customer_title.find_next_sibling('div', class_='registry-entry__body-href')
                        if not customer_value:
                            customer_value = customer_title.find_next_sibling('div', class_='registry-entry__body-value')
                        if customer_value:
                            customer = customer_value.text.strip()
                            if len(customer) > 100:
                                customer = customer[:97] + '...'
                    else:
                        body_href = block.find('div', class_='registry-entry__body-href')
                        if body_href:
                            customer = body_href.text.strip()
                            if len(customer) > 100:
                                customer = customer[:97] + '...'
                
                # federal law
                law_tag = block.find('div', class_='registry-entry__header-top__title')
                law = 'Не указан'
                if law_tag:
                    law_text = law_tag.text.strip()
                    if '44-ФЗ' in law_text:
                        law = '44-ФЗ'
                    elif '223-ФЗ' in law_text:
                        law = '223-ФЗ'
                    else:
                        law = law_text[:20]
                
                # price
                price = 'Не указана'
                price_block = block.find('div', class_='price-block')
                if price_block:
                    price_value = price_block.find('div', class_='price-block__value')
                    if price_value:
                        price_text = price_value.text.strip()
                        price_match = re.search(r'([\d\s]+[\.,]\d{2})', price_text)
                        if price_match:
                            price = price_match.group(1).replace(' ', '').replace(',', '.')
                        else:
                            price = price_text.replace('₽', '').replace('&8381;', '').strip()
                
                # end date
                end_date = 'Не указана'
                data_block = block.find('div', class_='data-block')
                if data_block:
                    titles = data_block.find_all('div', class_='data-block__title')
                    values = data_block.find_all('div', class_='data-block__value')
                    for i, title in enumerate(titles):
                        if 'окончание' in title.text.lower() or 'подачи' in title.text.lower():
                            if i < len(values):
                                end_date = values[i].text.strip()
                                break
                
                tenders[number] = {
                    'name': name,
                    'url': url,
                    'price': price,
                    'law': law,
                    'end_date': end_date,
                    'customer': customer,
                    'first_seen': current_date
                }
                
                # we note whether it is already in the database
                status = "New" if number not in self.existing_numbers else "Exists"
                print(f"{status} {number} | {law} | {price} rub.")
                
            except Exception as e:
                print(f"Block parsing error: {e}")
                continue
        
        return tenders
    
    # peeks at next page
    def has_next_page(self, soup):
        try:
            return soup.find('a', class_='paginator-button-next') is not None
        except:
            return False
    
    # chews through all pages and stashes along the way
    def parse_all_pages(self, max_pages=None):
        print("="*60)
        print("PARSING TENDERS (saving as we go)")
        print("="*60)
        
        page: int = 1
        total_new: int = 0
        
        while True:
            if max_pages and page > max_pages:
                print(f"Page limit reached ({max_pages})")
                break
            
            print(f"\n--- Page {page} ---")
            
            page_tenders, soup = self.parse_page(page)
            
            if not page_tenders:
                print("Tenders not found")
                break
            
            # we save only new tenders from this page
            new_on_page = self.save_tenders(page_tenders)
            total_new += new_on_page
            
            print(f"On page: new {new_on_page}, total added {total_new}")
            
            if not self.has_next_page(soup):
                print("Last page")
                break
            
            page += 1
            time.sleep(1)  # pause between requests
        
        print(f"\n TOTAL NEW TENDERS ADDED: {total_new}")
        print(f"Total in database: {len(self.existing_numbers)}")
        print(f"File: {self.excel_file}")
        
        return total_new

def Parse_gos_zakupki(interactive=True):
    parser = TenderParser('tenders.xlsx')

    if not interactive:
        try:
            parser.parse_all_pages()
        except Exception as e:
            print(f"Error: {e}")
            traceback.print_exc()
        return

    while True:
        print("\n" + "="*60)
        print("TENDER PARSER (saving as we go)")
        print("="*60)
        print("1.  Parse all pages")
        print("2.  Parse N pages")
        print("3.  Exit")
        print("="*60)
        
        choice = input("Choose an action: ").strip()
        
        if choice == '1':
            try:
                parser.parse_all_pages()
            except KeyboardInterrupt:
                print("\n\n Interrupted by user")
                print("Data already saved, can restart - duplicates will not be added")
            except Exception as e:
                print(f"Error: {e}")
                traceback.print_exc()
        
        elif choice == '2':
            try:
                pages = input("How many pages to parse? ").strip()
                max_pages = int(pages) if pages else 10
                parser.parse_all_pages(max_pages=max_pages)
            except KeyboardInterrupt:
                print("\n\n Interrupted by user")
            except Exception as e:
                print(f"Error: {e}")
        
        elif choice == '3':
            print("Goodbye!")
            break
        
        else:
            print("Invalid choice")

if __name__ == "__main__":
    


    Parse_gos_zakupki()