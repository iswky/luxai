from bs4 import BeautifulSoup
from bs4 import Tag
from urllib.parse import urljoin
from typing import List, Dict, Optional, Tuple, Any
import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import PatternFill
import os

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
    'Accept-Encoding': 'gzip, deflate, br',
    'Referer': 'https://zakupki.gov.ru/',
    'Connection': 'keep-alive',
}

green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")


def unarchive_file(file_path: str):
    """
    Безопасная распаковка архива с обработкой всех ошибок
    """

    import shutil
    from unzipall import extract
    from unzipall.exceptions import ArchiveExtractionError

    
    # Проверка существования файла
    if not os.path.exists(file_path):
        print(f"⚠️ Файл не найден: {file_path}")
        return False
    
    # Проверка, что это файл, а не папка
    if not os.path.isfile(file_path):
        print(f"⚠️ Путь не является файлом: {file_path}")
        return False
    
    try:
        # Пытаемся распаковать архив
        print(f"📦 Распаковка: {file_path}")
        extract(file_path)
        
        # Удаляем архив после успешной распаковки
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"✅ Архив удален: {file_path}")
        
    except ArchiveExtractionError as e:
        print(f"❌ Ошибка распаковки {file_path}: {e}")
        # Не удаляем архив, чтобы можно было попробовать позже
        return False
    except Exception as e:
        print(f"❌ Неожиданная ошибка при распаковке {file_path}: {e}")
        return False
    
    # Формируем пути для перемещения файлов
    try:
        # Получаем имя без расширения (более надежный способ)
        source_folder = os.path.splitext(file_path)[0]
        
        # Если папка не существует, возможно архив не создал папку
        if not os.path.exists(source_folder):
            print(f"⚠️ Папка {source_folder} не найдена после распаковки")
            return False
        
        # Родительская папка (на уровень выше)
        destination_folder = os.path.dirname(source_folder)
        
        print(f"📁 Исходная папка: {source_folder}")
        print(f"📂 Папка назначения: {destination_folder}")
        
        # Перемещаем все файлы и папки
        moved_count = 0
        error_count = 0
        
        for item in os.listdir(source_folder):
            source_path = os.path.join(source_folder, item)
            destination_path = os.path.join(destination_folder, item)
            
            try:
                # Если в папке назначения уже есть такой файл/папка
                if os.path.exists(destination_path):
                    # Генерируем уникальное имя
                    base, ext = os.path.splitext(item)
                    counter = 1
                    while os.path.exists(destination_path):
                        new_name = f"{base}_{counter}{ext}"
                        destination_path = os.path.join(destination_folder, new_name)
                        counter += 1
                    print(f"🔄 Переименован: {item} -> {os.path.basename(destination_path)}")
                
                # Перемещаем
                shutil.move(source_path, destination_path)
                moved_count += 1
                
            except Exception as e:
                print(f"❌ Ошибка перемещения {item}: {e}")
                error_count += 1
                continue
        
        # Удаляем исходную папку
        try:
            if os.path.exists(source_folder):
                os.rmdir(source_folder)  # Пробуем удалить пустую папку
                print(f"🗑️ Удалена папка: {source_folder}")
        except OSError:
            # Папка не пуста, удаляем рекурсивно
            try:
                shutil.rmtree(source_folder)
                print(f"🗑️ Удалена папка с содержимым: {source_folder}")
            except Exception as e:
                print(f"⚠️ Не удалось удалить папку {source_folder}: {e}")
        
        print(f"✅ Перемещено файлов: {moved_count}, ошибок: {error_count}")
        return True
        
    except Exception as e:
        print(f"❌ Ошибка при обработке распакованных файлов: {e}")
        return False

def download_file(url: str, filename: str = "file.html", tender_number: str = "unknown_tender/"):
    headers = {
        'User-Agent': 'Mozilla/5.0'
    }

    # Формируем путь: tenders_files/<tender_number>/
    path = os.path.join("tenders_files", tender_number)
    os.makedirs(path, exist_ok=True)
    filepath = os.path.join(path, filename)

    try:
        r = requests.get(url, headers=headers, stream=True, timeout=30)
        r.raise_for_status()
        
        with open(filepath, 'wb') as f:
            for chunk in r.iter_content(chunk_size=8192):
                f.write(chunk)

        if os.path.splitext(filepath)[1].lower() in ['.zip', '.rar', '.7z', '.tar']:
            unarchive_file(filepath)


        print('Файл', filename, 'успешно скачан')

        return filepath
        
    except requests.RequestException as e:
        print(f"❌ Ошибка скачивания {filename}: {e}")
        raise

def extract_file_links_44FZ(html_content: str, base_url: str = 'https://zakupki.gov.ru') -> List[Dict[str, Optional[str]]]:
    
    soup: BeautifulSoup = BeautifulSoup(html_content, 'html.parser')
    files: List[Dict[str, Optional[str]]] = []

    section: List = soup.find('div', class_= "blockFilesTabDocs")

    if not section:
        print("❌ Секция blockFilesTabDocs не найдена")
        return []
    
    attachments: List = section.find_all('div', class_="attachment row")
    print(f"📎 Найдено вложений: {len(attachments)}")

    for i, attachment in enumerate(attachments, 1):
        link: Optional[Tag] = attachment.find('a', href=lambda h: h and 'https://zakupki.gov.ru' in h)

        if link:
            file_url: str = link.get('href', '')
            file_title: str = link.get('title', '')
            file_name: str = link.text.strip()

            print(file_url)
            print(file_title)
            print('\n')
            
            file_info = {
                'url': file_url,
                'title': file_title
            }

            files.append(file_info)

    return files

def extract_file_links_223FZ(url: str, base_url: str = 'https://zakupki.gov.ru') -> List[Dict[str, Optional[str]]]:

    from playwright.sync_api import sync_playwright

    try:

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            
            page.goto(url)

            # ✅ Ждём внутри блока
            try:
                page.wait_for_selector("a[href*='download.html']", timeout=15000)
                html = page.content()
                browser.close()
                # Продолжаем обработку HTML
            except Exception:
                print(f"⚠️ Таймаут: элемент 'a[href*='download.html']' не найден на {url}")
                browser.close()
                return []

        # ✅ Используем html, который получили
        soup: BeautifulSoup = BeautifulSoup(html, 'html.parser')
        files: List[Dict[str, Optional[str]]] = []

        section: Tag = soup.find('section', class_="card-attachments")

        if not section:
            print("❌ Секция \"card-attachments\" не найдена")
            return []

        counts = section.find_all('span', class_=lambda x: x and 'count' in x.split())

        if not counts:
            print("❌ Секции \"count \" не найдена")
            return []

        for count in counts:
            link = count.find('a', string=lambda text: text and 'download.html')

            if link:

                from html import unescape

                file_url = "https://zakupki.gov.ru" + link.get('href')
                file_name: str = link.text.strip()

                tooltip = link.get('data-tooltip')
                tooltip_decoded = unescape(tooltip)
                file_name = BeautifulSoup(tooltip_decoded, 'html.parser').find('span', class_='custom-tooltiptext').text

                print(file_url)
                print(file_name)
                print('\n')
                
                file_info = {
                    'url': file_url,
                    'title': file_name
                }

                files.append(file_info)
            else:
                print("filelink not found")

        return files
    
    except Exception as e:
        print(f"Ошибка при загрузке {url}: {e}")
        return []

def get_link_to_file_page(tender_link: str) -> str:

    response = requests.get(tender_link, headers = headers)

    soup: BeautifulSoup = BeautifulSoup(response.text, 'html.parser')
    files: List[Dict[str, Optional[str]]] = []

    section: List = soup.find('div', class_= "tabsNav d-flex")

    if not section:
        print("❌ Секция tabsNav d-flex не найдена")
        return ""
    

    link: Optional[Tag] = section.find('a', href=lambda h: h and 'documents.html' in h)

    if link:
        docs_page_url: str = link.get('href', '')

        print(docs_page_url)
        print('\n')

        return docs_page_url

    return ""

def download_tenders_files():
    wb: Workbook = load_workbook("tenders.xlsx")
    ws: Worksheet = wb.active

    i = 2
    for row in range(2, ws.max_row + 1):
        tender_id: Any = ws.cell(row=row, column=1).value[2:]
        tender_link: str = ws.cell(row=row, column=3).value
        FZ: str = ws.cell(row=row, column=5).value
        Files_download_flag: Any = ws.cell(row=row, column=10).value

        if Files_download_flag == "True":
            continue

        url: str = "https://zakupki.gov.ru" + get_link_to_file_page(tender_link)

        print

        print(url)
        print(f"Сайт откуда скачиваем: {url}")
        print(f"скачиваем файлы для тендера: {tender_id}")

        if FZ == "44-ФЗ":
            # Получаем HTML
            response = requests.get(url, headers = headers)
            html_text: str = response.text
            files: List[Dict[str, Optional[str]]] = []
            files = extract_file_links_44FZ(html_text)
        elif FZ == "223-ФЗ":
            files = extract_file_links_223FZ(url)

        if not files:
            continue

        try:
            for file_info in files:
                download_file(file_info['url'], file_info['title'], tender_id)
            ws.cell(row=row, column=10).value = 'True'
            ws.cell(row=row, column=10).fill = green_fill
            wb.save("tenders.xlsx")
            
        except requests.RequestException as e:
            print()
        
        i += 1


if __name__ == "__main__":
    download_tenders_files()