from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl import load_workbook
from typing import List, Dict, Any
from bs4 import BeautifulSoup
import requests
import logging
import time
import re

logging.basicConfig(level = logging.DEBUG, filename = "parse_logs.log", filemode = "w")

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
    'Accept-Encoding': 'gzip, deflate, br',
    'Referer': 'https://zakupki.gov.ru/',
    'Connection': 'keep-alive',
}

def read_tenders_info(filename: str) -> List[Dict[str, Any]]:
    try:
        wb: Workbook = load_workbook(filename)
        ws: Worksheet = wb.active
        logging.debug(f"{filename} was opened to read information about tenders")

        pairs: List[Dict[str, Any]] = []
        for row in range(2, ws.max_row + 1):

            val_1: str = ws.cell(row=row, column=1).value[2:]
            val_9: str = ws.cell(row=row, column=9).value

            if val_9:
                continue

            tender_info = {
                'Number': val_1,
                'City': val_9,
                'row_num': row
            }
            pairs.append(tender_info)

        return pairs
    
    except FileNotFoundError as e:
        logging.error(f"Error in city_parse.py:read_tenders_info: File {filename} not found: {e}")
        raise FileNotFoundError("File {filename} not found: {e}")
        return []
        
    except PermissionError as e:
        print(f"Error in city_parse.py:read_tenders_info: No access to file {filename}: {e}")
        return []
        
    except KeyError as e:
        print(f"Error in city_parse.py:read_tenders_info: Error in data structures: {e}")
        return []
        
    except Exception as e:
        # Ловим все остальные непредвиденные ошибки
        print(f"Error in city_parse.py:read_tenders_info: Error in city_parse.py:read_tenders_info: {e}")
        print(f"Тип ошибки: {type(e).__name__}")
        return []
        



def extract_city_from_address(address: str) -> str:
    """Извлекает название города/населенного пункта из адреса"""
    
    # Паттерны для поиска города
    patterns = [
        r'г\.?\s+([А-Яа-я\-]+(?:\s+[А-Яа-я\-]+)?)',  # г. Москва, г Москва
        r'город\s+([А-Яа-я\-]+(?:\s+[А-Яа-я\-]+)?)',  # город Москва
        r'пос\.?\s+([А-Яа-я\-]+(?:\s+[А-Яа-я\-]+)?)',  # пос. Солнечный
        r'д\.?\s+([А-Яа-я\-]+(?:\s+[А-Яа-я\-]+)?)',    # д. Ивановка
        r'деревня\s+([А-Яа-я\-]+(?:\s+[А-Яа-я\-]+)?)', # деревня Ивановка
    ]
    
    for pattern in patterns:
        match = re.search(pattern, address, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    
    # Если нет маркеров, берем второй элемент по запятым
    parts = [p.strip() for p in address.split(',')]
    if len(parts) >= 2:
        # Пропускаем индекс (первый элемент)
        return parts[1]
    
    return parts[0] if parts else address

def extract_city(html_content: str) -> str:
    soup: BeautifulSoup = BeautifulSoup(html_content, 'html.parser')
    city: str

    sections: List = soup.find_all('section', class_= "blockInfo__section section")

    if not sections:
        print("❌ Секция blockInfo__section section не найдена")
        return "Не указан"
    
    print(f"Найдено секций: {len(sections)}")
    
    for section in sections:
        span = section.find('span', class_ = "section__title")
        
        if not span:
            continue

        if span.get_text(strip = True) != "Место поставки товара, выполнения работы или оказания услуги":
            continue

        span_dispatch = section.find('span', class_ = "section__info")
        if not span_dispatch:
            return "Не указан"

        place_of_dispatch: str = span_dispatch.get_text(strip = True)
        # print(place_of_dispatch)
            
        city: str = extract_city_from_address(place_of_dispatch)
        # print(city)

        return city

    return "Не указан"

def city_parse():
    try:
        tenders_info: List[Dict[str, Any]] = read_tenders_info("tenders.xlsx")

        wb: Workbook = load_workbook("tenders.xlsx")
        ws: Worksheet = wb.active

        for tender in tenders_info:
            tender_num: str = tender['Number']
            city: str = tender['City']
            row_num: int = tender['row_num']
            url = 'https://zakupki.gov.ru/epz/order/notice/ea20/view/common-info.html?regNumber=' + tender_num

            response = requests.get(url, headers = headers)

            # Получаем HTML как строку
            html_text: str = response.text

            city = extract_city(html_text)

            ws.cell(row = row_num, column = 9).value = city
            wb.save("tenders.xlsx")

            time.sleep(1)
    except:
        print()