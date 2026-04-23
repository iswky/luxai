from bs4 import BeautifulSoup
from urllib.parse import urljoin
from typing import List, Dict, Optional, Tuple, Any
import requests
import os
# import re
# import zipfile
# from urllib.parse import urljoin, urlparse, parse_qs
# import time
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import PatternFill
# import mimetypes
# import cgi

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
    'Accept-Encoding': 'gzip, deflate, br',
    'Referer': 'https://zakupki.gov.ru/',
    'Connection': 'keep-alive',
}

green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

def download_file(url: str, filename: str = "file.html", tender_number: str = "unknown_tender/"):
    headers = {
        'User-Agent': 'Mozilla/5.0'
    }

    # Формируем путь: tenders_files/<tender_number>/
    path = os.path.join("tenders_files", tender_number)
    os.makedirs(path, exist_ok=True)
    filepath = os.path.join(path, filename)

    try:
        r = requests.get(url, headers=headers, stream=True, timeout=30)
        r.raise_for_status()
        
        with open(filepath, 'wb') as f:
            for chunk in r.iter_content(chunk_size=8192):
                f.write(chunk)

        print('Файл', filename, 'успешно скачан')

        return filepath
        
    except requests.RequestException as e:
        print(f"❌ Ошибка скачивания {filename}: {e}")
        raise

def extract_file_links(html_content: str, base_url: str = 'https://zakupki.gov.ru') -> List[Dict[str, Optional[str]]]:
    
    soup: BeautifulSoup = BeautifulSoup(html_content, 'html.parser')
    files: List[Dict[str, Optional[str]]] = []

    section: List = soup.find('div', class_= "blockFilesTabDocs")

    if not section:
        print("❌ Секция blockFilesTabDocs не найдена")
        return []
    
    attachments: List = section.find_all('div', class_="attachment row")
    print(f"📎 Найдено вложений: {len(attachments)}")

    for i, attachment in enumerate(attachments, 1):
        link: Optional[Tag] = attachment.find('a', href=lambda h: h and 'https://zakupki.gov.ru' in h)

        if link:
            file_url: str = link.get('href', '')
            file_title: str = link.get('title', '')
            file_name: str = link.text.strip()

            print(file_url)
            print(file_title)
            print('\n')
            
            file_info = {
                'url': file_url,
                'title': file_title
            }

            files.append(file_info)

    return files

def read_tenders_info(filename: str) -> List[Tuple[Any, Any]]:
    wb: Workbook = load_workbook(filename)
    ws: Worksheet = wb.active

    pairs: List[Tuple[Any, Any]] = []
    for row in range(2, ws.max_row + 1):
        val_1: Any = ws.cell(row=row, column=1).value[2:]
        val_10: Any = ws.cell(row=row, column=10).value
        pairs.append((val_1, val_10))

    return pairs

def main():
    
    tenders_info: List[Tuple[Any, Any]] = read_tenders_info("tenders.xlsx")

    wb: Workbook = load_workbook("tenders.xlsx")
    ws: Worksheet = wb.active

    i = 2
    for number, flag in tenders_info:
        url = 'https://zakupki.gov.ru/epz/order/notice/ea20/view/documents.html?regNumber=' + number

        print("сайт откуда скачиваем: ", url)

        if (flag == 'False'):
            print("скачиваем файлы для тендера:", number)
            response = requests.get(url, headers = headers)

            # Получаем HTML как строку
            html_text: str = response.text

            files: List[Dict[str, Optional[str]]] = extract_file_links(html_text)

            try:
                for file_info in files:
                    download_file(file_info['url'], file_info['title'], number)
                ws.cell(row=i, column=10).value = 'True'
                ws.cell(row=i, column=10).fill = green_fill
                wb.save("tenders.xlsx")
                
            except requests.RequestException as e:
                print()
        
        i += 1



if __name__ == "__main__":
    main()

