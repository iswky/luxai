import requests
from bs4 import BeautifulSoup
import csv
from datetime import datetime
import os
import time
import traceback
import re

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
    print("✅ openpyxl загружен успешно")
except ImportError as e:
    OPENPYXL_AVAILABLE = False
    print(f"❌ openpyxl не установлен: {e}")

class TenderParser:
    """Парсер тендеров с расширенными данными"""
    
    def __init__(self, excel_file='tenders.xlsx'):
        self.excel_file = excel_file
        self.csv_file = excel_file.replace('.xlsx', '.csv')
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
        }
        
        self.base_url = 'https://zakupki.gov.ru/epz/order/extendedsearch/results.html'
        
        self.base_params = {
            'morphology': 'on',
            'search-filter': 'Дате размещения',
            'sortDirection': 'false',
            'recordsPerPage': '_10',
            'showLotsInfoHidden': 'false',
            'sortBy': 'PUBLISH_DATE',
            'fz44': 'on',
            'fz223': 'on',
            'af': 'on',
            'currencyIdGeneral': '-1',
            'okpd2IdsWithNested': 'on',
            'okpd2Ids': '8873938,8874157,8874056,8873907,8874087,8874160,8874159,8874058,8874059,8874055,8874158,8873937,8873908,8873906,8874054,8874060,8874061',
            'okpd2IdsCodes': '63,61.1,26.3,27,32.9,61.9,61.3,26.5,26.6,26.2,61.2,62,28,26,26.1,26.7,26.8'
        }
        
        print("\n" + "="*60)
        print("🎯 ПАРСЕР ЗАПУЩЕН")
        print("="*60)
    
    def load_existing_tenders(self):
        """Загружает существующие номера тендеров из Excel"""
        existing_numbers = set()
        
        if OPENPYXL_AVAILABLE and os.path.exists(self.excel_file):
            try:
                wb = load_workbook(self.excel_file)
                ws = wb.active
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        existing_numbers.add(str(row[0]).strip())
                
                print(f"📂 Загружено существующих тендеров: {len(existing_numbers)}")
            except Exception as e:
                print(f"Ошибка при загрузке: {e}")
        else:
            print("📂 Файл не найден или пуст")
        
        return existing_numbers
    
    def save_tenders(self, new_tenders):
        """Сохраняет новые тендеры"""
        if not new_tenders:
            return 0
        
        if OPENPYXL_AVAILABLE:
            return self._append_to_excel(new_tenders)
        else:
            return self._append_to_csv(new_tenders)
    
    def _append_to_excel(self, new_tenders):
        """Добавляет в Excel"""
        try:
            if os.path.exists(self.excel_file):
                wb = load_workbook(self.excel_file)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Тендеры"
                # Расширенные заголовки с заказчиком
                headers = ['Номер тендера', 'Название/Объект закупки', 'Ссылка', 
                          'Цена (руб)', 'ФЗ', 'Окончание подачи заявок', 'Заказчик', 'Дата добавления']
                ws.append(headers)
                
                header_font = Font(bold=True, size=11, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 70
                ws.column_dimensions['C'].width = 60
                ws.column_dimensions['D'].width = 18
                ws.column_dimensions['E'].width = 10
                ws.column_dimensions['F'].width = 20
                ws.column_dimensions['G'].width = 50
                ws.column_dimensions['H'].width = 20
            
            next_row = ws.max_row + 1
            for number, info in new_tenders.items():
                ws.cell(row=next_row, column=1, value=number)
                ws.cell(row=next_row, column=2, value=info['name'])
                ws.cell(row=next_row, column=3, value=info['url'])
                ws.cell(row=next_row, column=4, value=info.get('price', 'Не указана'))
                ws.cell(row=next_row, column=5, value=info.get('law', 'Не указан'))
                ws.cell(row=next_row, column=6, value=info.get('end_date', 'Не указана'))
                ws.cell(row=next_row, column=7, value=info.get('customer', 'Не указан'))
                ws.cell(row=next_row, column=8, value=info['first_seen'])
                next_row += 1
            
            wb.save(self.excel_file)
            print(f"✅ СОХРАНЕНО {len(new_tenders)} тендеров в {self.excel_file}")
            return len(new_tenders)
            
        except Exception as e:
            print(f"❌ Ошибка сохранения: {e}")
            traceback.print_exc()
            return 0
    
    def _append_to_csv(self, new_tenders):
        """Сохраняет в CSV"""
        try:
            file_exists = os.path.exists(self.csv_file)
            
            with open(self.csv_file, 'a', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                if not file_exists:
                    writer.writerow(['number', 'name', 'url', 'price', 'law', 'end_date', 'customer', 'first_seen'])
                
                for number, info in new_tenders.items():
                    writer.writerow([
                        number, 
                        info['name'], 
                        info['url'],
                        info.get('price', 'Не указана'),
                        info.get('law', 'Не указан'),
                        info.get('end_date', 'Не указана'),
                        info.get('customer', 'Не указан'),
                        info['first_seen']
                    ])
            
            print(f"✅ СОХРАНЕНО {len(new_tenders)} тендеров в {self.csv_file}")
            return len(new_tenders)
        except Exception as e:
            print(f"❌ Ошибка: {e}")
            return 0
    
    def parse_page(self, page_number=1):
        """Парсит страницу"""
        params = self.base_params.copy()
        params['pageNumber'] = str(page_number)
        
        try:
            print(f"📥 Загрузка страницы {page_number}...")
            response = requests.get(self.base_url, headers=self.headers, params=params, timeout=15)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            return self.extract_tenders(soup), soup
        except Exception as e:
            print(f"❌ Ошибка страницы {page_number}: {e}")
            return [], None
    
    def extract_tenders(self, soup):
        """Извлекает тендеры с расширенными данными (включая заказчика)"""
        tenders = {}
        current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        blocks = soup.find_all('div', class_='search-registry-entry-block')
        print(f"🔍 Найдено блоков: {len(blocks)}")
        
        for block in blocks:
            try:
                # Номер и ссылка
                number_div = block.find('div', class_='registry-entry__header-mid__number')
                if not number_div:
                    continue
                
                link = number_div.find('a')
                if not link:
                    continue
                
                number = link.text.strip()
                url = 'https://zakupki.gov.ru' + link.get('href')
                
                # Название
                name_div = block.find('div', class_='registry-entry__body-value')
                name = name_div.text.strip() if name_div else 'Название не указано'
                if len(name) > 200:
                    name = name[:197] + '...'
                
                # ЗАКАЗЧИК
                customer = 'Не указан'
                customer_block = block.find('div', class_='registry-entry__body-block')
                if customer_block:
                    # Ищем блок с заголовком "Заказчик"
                    customer_title = customer_block.find('div', class_='registry-entry__body-title', string=re.compile(r'Заказчик|Организация'))
                    if customer_title:
                        # Берем следующий элемент с классом body-href или body-value
                        customer_value = customer_title.find_next_sibling('div', class_='registry-entry__body-href')
                        if not customer_value:
                            customer_value = customer_title.find_next_sibling('div', class_='registry-entry__body-value')
                        if customer_value:
                            customer = customer_value.text.strip()
                            # Ограничиваем длину
                            if len(customer) > 100:
                                customer = customer[:97] + '...'
                    else:
                        # Альтернативный поиск
                        body_href = block.find('div', class_='registry-entry__body-href')
                        if body_href:
                            customer = body_href.text.strip()
                            if len(customer) > 100:
                                customer = customer[:97] + '...'
                
                # ФЗ (закон)
                law_tag = block.find('div', class_='registry-entry__header-top__title')
                law = 'Не указан'
                if law_tag:
                    law_text = law_tag.text.strip()
                    if '44-ФЗ' in law_text:
                        law = '44-ФЗ'
                    elif '223-ФЗ' in law_text:
                        law = '223-ФЗ'
                    else:
                        law = law_text[:20]
                
                # Цена
                price = 'Не указана'
                price_block = block.find('div', class_='price-block')
                if price_block:
                    price_value = price_block.find('div', class_='price-block__value')
                    if price_value:
                        price_text = price_value.text.strip()
                        price_match = re.search(r'([\d\s]+[\.,]\d{2})', price_text)
                        if price_match:
                            price = price_match.group(1).replace(' ', '').replace(',', '.')
                        else:
                            price = price_text.replace('₽', '').replace('&8381;', '').strip()
                
                # Дата окончания подачи заявок
                end_date = 'Не указана'
                data_block = block.find('div', class_='data-block')
                if data_block:
                    titles = data_block.find_all('div', class_='data-block__title')
                    values = data_block.find_all('div', class_='data-block__value')
                    for i, title in enumerate(titles):
                        if 'окончание' in title.text.lower() or 'подачи' in title.text.lower():
                            if i < len(values):
                                end_date = values[i].text.strip()
                                break
                
                tenders[number] = {
                    'name': name,
                    'url': url,
                    'price': price,
                    'law': law,
                    'end_date': end_date,
                    'customer': customer,
                    'first_seen': current_date
                }
                
                # Короткий вывод в консоль
                customer_short = customer[:30] + '...' if len(customer) > 30 else customer
                print(f"  📝 {number} | {law} | {price} руб. | {customer_short}")
                
            except Exception as e:
                print(f"  ⚠️ Ошибка парсинга блока: {e}")
                continue
        
        return tenders
    
    def has_next_page(self, soup):
        """Проверяет следующую страницу"""
        try:
            return soup.find('a', class_='paginator-button-next') is not None
        except:
            return False
    
    def parse_all_pages(self, max_pages=None):
        """Парсит все страницы и сохраняет"""
        print("="*60)
        print("🚀 ПАРСИНГ ТЕНДЕРОВ (с ценой, ФЗ, датой окончания, заказчиком)")
        print("="*60)
        
        all_tenders = {}
        page = 1
        total_tenders_found = 0
        
        while True:
            if max_pages and page > max_pages:
                print(f"🏁 Достигнут лимит страниц ({max_pages})")
                break
            
            print(f"\n--- Страница {page} ---")
            
            page_tenders, soup = self.parse_page(page)
            
            if not page_tenders:
                print("❌ Тендеры не найдены")
                break
            
            # Добавляем все тендеры
            new_in_page = 0
            for number, info in page_tenders.items():
                if number not in all_tenders:
                    all_tenders[number] = info
                    new_in_page += 1
            
            total_tenders_found += new_in_page
            print(f"📊 На странице: новых {new_in_page}, всего собрано {len(all_tenders)}")
            
            if not self.has_next_page(soup):
                print("🏁 Последняя страница")
                break
            
            page += 1
            time.sleep(1)
        
        print(f"\n📊 ВСЕГО НАЙДЕНО ТЕНДЕРОВ: {len(all_tenders)}")
        
        if all_tenders:
            result = self.save_tenders(all_tenders)
            print(f"\n✅ СОХРАНЕНО: {result} тендеров")
            
            # Показываем пример первых 5
            print("\n📋 ПРИМЕР НАЙДЕННЫХ ДАННЫХ:")
            print("-" * 90)
            for i, (number, info) in enumerate(list(all_tenders.items())[:5], 1):
                print(f"{i}. {number}")
                print(f"   Закон: {info['law']}")
                print(f"   Цена: {info['price']} руб.")
                print(f"   Окончание: {info['end_date']}")
                print(f"   Заказчик: {info['customer'][:60]}...")
                print(f"   Название: {info['name'][:60]}...")
                print()
        else:
            print("❌ НИЧЕГО НЕ НАЙДЕНО")
        
        return all_tenders


def main():
    parser = TenderParser('tenders.xlsx')
    
    while True:
        print("\n" + "="*60)
        print("ПАРСЕР ТЕНДЕРОВ (расширенный)")
        print("="*60)
        print("1. 🚀 Парсить все страницы")
        print("2. 🔢 Парсить N страниц")
        print("3. 🚪 Выход")
        print("="*60)
        
        choice = input("Выберите действие: ").strip()
        
        if choice == '1':
            try:
                parser.parse_all_pages()
            except KeyboardInterrupt:
                print("\n\n⚠️ Прервано")
            except Exception as e:
                print(f"❌ Ошибка: {e}")
                traceback.print_exc()
        
        elif choice == '2':
            try:
                pages = input("Сколько страниц парсить? ").strip()
                max_pages = int(pages) if pages else 10
                parser.parse_all_pages(max_pages=max_pages)
            except KeyboardInterrupt:
                print("\n\n⚠️ Прервано")
            except Exception as e:
                print(f"❌ Ошибка: {e}")
        
        elif choice == '3':
            print("👋 До свидания!")
            break
        
        else:
            print("❌ Неверный выбор")

if __name__ == "__main__":
    main()